"""
report.py
---------
Writes a multi-sheet, formatted Excel summary report suitable for
stakeholder review -- the "tell the story behind the numbers" output
referenced in the PRD.

Sheets:
  1. Summary       - run metadata + headline counts
  2. Issue Log     - the full auditable issue log
  3. Reconciliation - full record-level reconciliation detail
  4. Limit Monitoring - exposure vs. limit by counterparty/asset_class
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

ACCENT = "1F4E5F"
ACCENT_LIGHT = "DCE9EE"
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, color=ACCENT, size=16)
SUBTITLE_FONT = Font(name="Arial", italic=True, color="595959", size=10)
LABEL_FONT = Font(name="Arial", bold=True, size=11)
BODY_FONT = Font(name="Arial", size=10)
BREACH_FILL = PatternFill("solid", start_color="FBE2E2")
HEADER_FILL = PatternFill("solid", start_color=ACCENT)
thin = Side(style="thin", color="CCCCCC")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_df(ws, df, start_row=1, highlight_col=None, highlight_val=None):
    for j, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=j, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=1):
            val = row[col]
            cell = ws.cell(row=i, column=j, value=None if (val is None or val != val) else val)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            if highlight_col and highlight_val is not None and col == highlight_col and val == highlight_val:
                for k in range(1, len(df.columns) + 1):
                    ws.cell(row=i, column=k).fill = BREACH_FILL

    for j, col in enumerate(df.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].astype(str)])
        ws.column_dimensions[get_column_letter(j)].width = min(max(max_len + 2, 12), 55)

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)


def write_excel_report(report_path, reconciled, breaks_only, limit_results, issue_log_df,
                        run_date, fo_count, rs_count):
    wb = Workbook()

    # --- Sheet 1: Summary ---
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Risk Data Reconciliation & Limit Monitoring Report"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"Run date: {run_date}  |  Tolerance threshold: 1%"
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")

    clean = (reconciled["break_category"] == "clean_match").sum()
    breach_count = int(limit_results["is_breach"].sum())
    summary_rows = [
        ("Front Office records ingested", fo_count),
        ("Risk System records ingested", rs_count),
        ("Total reconciled records", len(reconciled)),
        ("Clean matches", int(clean)),
        ("Reconciliation breaks", len(breaks_only)),
        ("Limit breaches", breach_count),
        ("Open issues logged", len(issue_log_df)),
    ]
    r = 4
    for label, value in summary_rows:
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        ws.cell(row=r, column=2, value=value).font = BODY_FONT
        r += 1

    r += 1
    ws.cell(row=r, column=1, value="Break Type Breakdown").font = LABEL_FONT
    r += 1
    for cat, n in reconciled["break_category"].value_counts().items():
        ws.cell(row=r, column=1, value=cat).font = BODY_FONT
        ws.cell(row=r, column=2, value=int(n)).font = BODY_FONT
        r += 1

    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 16

    # --- Sheet 2: Issue Log ---
    ws2 = wb.create_sheet("Issue Log")
    _write_df(ws2, issue_log_df)

    # --- Sheet 3: Reconciliation Detail ---
    recon_display = reconciled[[
        "trade_id", "counterparty_fo", "counterparty_rs", "asset_class_fo",
        "notional_fo", "notional_rs", "pct_diff", "currency_fo", "currency_rs",
        "break_category",
    ]].rename(columns={
        "counterparty_fo": "counterparty (FO)", "counterparty_rs": "counterparty (RS)",
        "asset_class_fo": "asset_class", "notional_fo": "notional_fo_usd",
        "notional_rs": "notional_rs_usd", "currency_fo": "ccy_fo", "currency_rs": "ccy_rs",
    })
    ws3 = wb.create_sheet("Reconciliation Detail")
    _write_df(ws3, recon_display, highlight_col="break_category")
    for i in range(2, len(recon_display) + 2):
        if ws3.cell(row=i, column=recon_display.columns.get_loc("break_category") + 1).value != "clean_match":
            for k in range(1, len(recon_display.columns) + 1):
                ws3.cell(row=i, column=k).fill = BREACH_FILL

    # --- Sheet 4: Limit Monitoring ---
    limit_display = limit_results.rename(columns={
        "aggregate_exposure_usd": "aggregate_exposure_usd",
        "max_notional_usd": "limit_usd",
    })
    ws4 = wb.create_sheet("Limit Monitoring")
    _write_df(ws4, limit_display, highlight_col="is_breach", highlight_val=True)

    wb.save(report_path)
